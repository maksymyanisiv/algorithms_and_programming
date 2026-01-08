#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Скрипт для виконання лабораторної роботи №7
"Робота з базами даних в табличному процесорі"
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import random

def create_lab_work_7():
    wb = Workbook()
    
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    create_salary_sheet(wb)
    create_salary_formulas_sheet(wb)
    create_statistics_sheet(wb)
    create_autofilter_sheet(wb)
    create_advanced_filter_sheet(wb)
    
    filename = "Yanisiv.xlsx"
    wb.save(filename)
    print(f"Файл {filename} успішно створено!")
    return filename

def create_salary_sheet(wb):
    ws = wb.create_sheet("Зарплата", 0)
    
    ws.sheet_properties.tabColor = "FF6B9D"  # Рожевий колір ярлика
    
    ws['A1'] = "Зарплата працівників в ІТ"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    headers = ['Порядковий номер', 'Ідентифікаційний код працівника', 
               'Прізвище та ім\'я', 'Нарахована зарплата (грн.)', 
               'Податок із зарплати (грн.)', 'До видачі працівнику (грн.)']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    random.seed(42)
    employees = [
        ("Іваненко Іван", "1234567890", 7500),
        ("Петренко Петро", "2345678901", 12000),
        ("Коваленко Олена", "3456789012", 25000),
        ("Мельник Андрій", "4567890123", 15000),
        ("Шевченко Марія", "5678901234", 8000),
        ("Бондаренко Олександр", "6789012345", 30000),
        ("Ткаченко Наталія", "7890123456", 9500),
        ("Морозенко Дмитро", "8901234567", 18000),
        ("Лисенко Катерина", "9012345678", 22000),
        ("Гриценко Сергій", "0123456789", 11000),
        ("Савченко Оксана", "1122334455", 16000),
        ("Романенко Віктор", "2233445566", 19000)
    ]
    
    row = 4
    for idx, (name, code, salary) in enumerate(employees, start=1):
        ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal='center')
        
        ws.cell(row=row, column=2, value=code).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2).number_format = '@'
        
        ws.cell(row=row, column=3, value=name)
        
        ws.cell(row=row, column=4, value=salary)
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        
        tax_formula = f'=IF(D{row}<=8000, D{row}*0.1, IF(D{row}>20000, D{row}*0.2, D{row}*0.15))'
        ws.cell(row=row, column=5, value=tax_formula)
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        
        payout_formula = f'=D{row}-E{row}'
        ws.cell(row=row, column=6, value=payout_formula)
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row += 1
    
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 28
    ws.row_dimensions[3].height = 40

def create_salary_formulas_sheet(wb):
    source_ws = wb['Зарплата']
    ws = wb.copy_worksheet(source_ws)
    ws.title = "Зарплата (формули)"
    ws.sheet_view.showFormulas = True

def create_statistics_sheet(wb):
    ws = wb.create_sheet("Статистика ІТ")
    
    ws['A1'] = "Статистика ІТ-галузі в Україні"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    headers = ['Область', 'Кількість ФОП-чоловіки', 'Кількість ФОП-жінки', 
               'Відсоток ФОП-чоловіки (%)', 'Сума сплачених податків ФОП-ІТ (грн.)']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    random.seed(100)
    regions_data = [
        ("Київська", 1250, 450, 500000),
        ("Харківська", 890, 320, 380000),
        ("Одеська", 750, 280, 320000),
        ("Дніпропетровська", 680, 250, 290000),
        ("Львівська", 920, 340, 420000),
        ("Чернігівська", 420, 180, 180000),
        ("Черкаська", 380, 150, 160000),
        ("Чернівецька", 350, 140, 150000),
        ("Вінницька", 450, 190, 200000),
        ("Запорізька", 520, 200, 220000),
        ("Полтавська", 480, 180, 200000),
        ("Сумська", 320, 130, 140000),
        ("Тернопільська", 410, 160, 175000),
        ("Хмельницька", 390, 150, 165000),
        ("Житомирська", 360, 140, 155000)
    ]
    
    row = 4
    for region, men, women, taxes in regions_data:
        ws.cell(row=row, column=1, value=region)
        
        ws.cell(row=row, column=2, value=men)
        ws.cell(row=row, column=2).number_format = '#,##0'
        
        ws.cell(row=row, column=3, value=women)
        ws.cell(row=row, column=3).number_format = '#,##0'
        
        percent_formula = f'=B{row}/(B{row}+C{row})*100'
        ws.cell(row=row, column=4, value=percent_formula)
        ws.cell(row=row, column=4).number_format = '0.00'
        
        ws.cell(row=row, column=5, value=taxes)
        ws.cell(row=row, column=5).number_format = '#,##0'
        
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row += 1
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 35
    ws.row_dimensions[3].height = 50
    
    ws.auto_filter.ref = f"A3:E{row-1}"

def create_autofilter_sheet(wb):
    source_ws = wb['Статистика ІТ']
    ws = wb.copy_worksheet(source_ws)
    ws.title = "Автофільтр"
    
    ws.auto_filter.ref = f"A3:E{ws.max_row}"

def create_advanced_filter_sheet(wb):
    ws = wb.create_sheet("Розширений фільтр")
    
    ws['A1'] = "Діяльність працівників фірми ІТ-Market@"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    headers = ['Прізвище', 'Ім\'я', 'Посада', 'Стаж роботи (роки)', 
               'Внесок у розвиток фірми ($)', 'Оцінка роботи']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    random.seed(200)
    employees_data = [
        ("Іваненко", "Іван", "Розробник", 5, 850, "Відмінно"),
        ("Петренко", "Петро", "Тестувальник", 3, 600, "Добре"),
        ("Коваленко", "Олена", "Дизайнер", 4, 750, "Відмінно"),
        ("Мельник", "Андрій", "Розробник", 6, 950, "Відмінно"),
        ("Шевченко", "Марія", "Менеджер", 2, 500, "Добре"),
        ("Бондаренко", "Олександр", "Розробник", 7, 1100, "Відмінно"),
        ("Ткаченко", "Наталія", "Тестувальник", 4, 650, "Добре"),
        ("Морозенко", "Дмитро", "Розробник", 5, 800, "Відмінно"),
        ("Лисенко", "Катерина", "Дизайнер", 3, 550, "Добре"),
        ("Гриценко", "Сергій", "Розробник", 4, 720, "Добре"),
        ("Савченко", "Оксана", "Менеджер", 3, 580, "Добре"),
        ("Романенко", "Віктор", "Розробник", 6, 900, "Відмінно"),
        ("Василенко", "Олег", "Тестувальник", 5, 700, "Добре"),
        ("Білоус", "Тетяна", "Дизайнер", 4, 680, "Добре"),
        ("Петренко", "Максим", "Розробник", 3, 620, "Добре"),
        ("Бондар", "Юлія", "Менеджер", 2, 480, "Задовільно"),
        ("Петренко", "Ірина", "Тестувальник", 4, 660, "Добре"),
        ("Біленко", "Олексій", "Розробник", 5, 820, "Відмінно"),
        ("Петренко", "Анна", "Дизайнер", 3, 590, "Добре"),
        ("Бойко", "Володимир", "Розробник", 7, 1050, "Відмінно"),
        ("Іванов", "Іван", "Тестувальник", 4, 680, "Добре"),
        ("Петров", "Петро", "Дизайнер", 3, 560, "Добре"),
        ("Іващенко", "Іван", "Менеджер", 2, 520, "Добре"),
        ("Петрушенко", "Петро", "Розробник", 5, 780, "Відмінно"),
        ("Іванчук", "Іван", "Розробник", 6, 920, "Відмінно")
    ]
    
    row = 4
    for surname, name, position, experience, contribution, rating in employees_data:
        ws.cell(row=row, column=1, value=surname)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=position)
        ws.cell(row=row, column=4, value=experience)
        ws.cell(row=row, column=5, value=contribution)
        ws.cell(row=row, column=5).number_format = '#,##0'
        ws.cell(row=row, column=6, value=rating)
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row += 1
    
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 18
    ws.row_dimensions[3].height = 50
    
    ws.auto_filter.ref = f"A3:F{row-1}"

if __name__ == "__main__":
    create_lab_work_7()

