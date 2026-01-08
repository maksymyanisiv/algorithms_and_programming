#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Скрипт для створення Excel файлу з лабораторною роботою з табличних процесорів
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import math

def create_spreadsheet():
    wb = Workbook()
    
    # Видаляємо дефолтний лист
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 1. Лист "Видатки"
    create_vydatky_sheet(wb)
    
    # 1.1. Копія таблиці з формулами (пункт 6)
    create_vydatky_formulas_sheet(wb)
    
    # 2. Лист "Видатки на місяць"
    create_monthly_expenses_sheet(wb)
    
    # 3. Лист "Прайс"
    create_price_sheet(wb)
    
    # 4. Лист "Функція"
    create_function_sheet(wb)
    
    # Зберігаємо файл
    filename = "Yanisiv.xlsx"
    wb.save(filename)
    print(f"Файл {filename} успішно створено!")
    return filename

def create_vydatky_sheet(wb):
    ws = wb.create_sheet("Видатки", 0)
    
    # Заголовок
    ws['A1'] = "ВИДАТКИ НА РІК"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Заголовки таблиці
    headers = ['Вид видатків', 'Січень', 'Лютий', 'Березень', 'Квітень', 'Травень', 
               'Червень', 'Липень', 'Серпень', 'Вересень', 'Жовтень', 'Листопад', 'Грудень', 'Сума за рік']
    
    months = ['Січень', 'Лютий', 'Березень', 'Квітень', 'Травень', 
              'Червень', 'Липень', 'Серпень', 'Вересень', 'Жовтень', 'Листопад', 'Грудень']
    
    # Записуємо заголовки
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=45)
        # Без кольорів - базовий вигляд
        # cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Види видатків
    expense_types = ['Продукти', 'Транспорт', 'Комунальні послуги', 'Розваги', 'Одяг', 'Медицина']
    
    # Заповнюємо дані
    import random
    random.seed(42)  # Для відтворюваності
    
    for row_idx, expense_type in enumerate(expense_types, start=4):
        ws.cell(row=row_idx, column=1, value=expense_type).font = Font(bold=True)
        
        # Генеруємо випадкові суми для кожного місяця
        monthly_amounts = []
        for month in range(12):
            amount = random.randint(500, 5000)
            monthly_amounts.append(amount)
            cell = ws.cell(row=row_idx, column=month + 2, value=amount)
            cell.number_format = '#,##0'
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Сума за рік
        sum_formula = f"=SUM(B{row_idx}:M{row_idx})"
        ws.cell(row=row_idx, column=14, value=sum_formula)
        ws.cell(row=row_idx, column=14).number_format = '#,##0'
        ws.cell(row=row_idx, column=14).border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Рядок з мінімальними значеннями
    row_idx = 10
    ws.cell(row=row_idx, column=1, value="Мінімум").font = Font(bold=True)
    for month in range(12):
        col = month + 2
        col_letter = get_column_letter(col)
        min_formula = f"=MIN({col_letter}4:{col_letter}9)"
        ws.cell(row=row_idx, column=col, value=min_formula)
        ws.cell(row=row_idx, column=col).number_format = '#,##0'
        ws.cell(row=row_idx, column=col).border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Рядок з максимальними значеннями
    row_idx = 11
    ws.cell(row=row_idx, column=1, value="Максимум").font = Font(bold=True)
    for month in range(12):
        col = month + 2
        col_letter = get_column_letter(col)
        max_formula = f"=MAX({col_letter}4:{col_letter}9)"
        ws.cell(row=row_idx, column=col, value=max_formula)
        ws.cell(row=row_idx, column=col).number_format = '#,##0'
        ws.cell(row=row_idx, column=col).border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Рядок з середніми значеннями
    row_idx = 12
    ws.cell(row=row_idx, column=1, value="Середнє").font = Font(bold=True)
    for month in range(12):
        col = month + 2
        col_letter = get_column_letter(col)
        avg_formula = f"=AVERAGE({col_letter}4:{col_letter}9)"
        ws.cell(row=row_idx, column=col, value=avg_formula)
        ws.cell(row=row_idx, column=col).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=col).border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Налаштування ширини стовпців
    ws.column_dimensions['A'].width = 20
    for col in range(2, 15):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.row_dimensions[3].height = 60
    
    # Переконуємося, що режим показу формул вимкнено на основному листі
    ws.sheet_view.showFormulas = False
    
    # Створюємо допоміжний вертикальний стовпець з даними для графіка
    # Це потрібно, щоб openpyxl правильно інтерпретував дані як одну серію
    chart_data_start_row = 15
    chart_data_col = 15  # Стовпець O
    
    # Копіюємо формули з рядка середніх значень вертикально
    for idx, col in enumerate(range(2, 14), start=0):  # Стовпці B-M
        source_col_letter = get_column_letter(col)
        target_row = chart_data_start_row + idx
        # Створюємо формулу, яка посилається на комірку в рядку 12
        formula = f"={source_col_letter}12"
        ws.cell(row=target_row, column=chart_data_col, value=formula)
    
    # Створюємо рядок з назвами місяців для категорій графіка (без обертання тексту)
    chart_row = 14
    for col, month in enumerate(months, start=2):
        ws.cell(row=chart_row, column=col, value=month)
    
    # Графік середніх значень
    chart = LineChart()
    chart.title = "Графік видатків за середнім значенням"
    chart.style = 13
    chart.y_axis.title = 'Сума (грн)'
    chart.x_axis.title = 'Місяць'
    
    # Дані з вертикального стовпця (одна серія з 12 точками)
    data = Reference(ws, min_col=chart_data_col, min_row=chart_data_start_row, 
                     max_col=chart_data_col, max_row=chart_data_start_row + 11)
    
    # Категорії (місяці) з рядка без обертання
    cats = Reference(ws, min_col=2, min_row=chart_row, max_col=13, max_row=chart_row)
    
    # Додаємо дані - вертикальний стовпець = одна серія
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    
    ws.add_chart(chart, "A15")

def create_vydatky_formulas_sheet(wb):
    """Створює копію таблиці 'Видатки' з режимом відображення формул та зміненим форматом чисел"""
    source_ws = wb['Видатки']
    
    # Використовуємо copy_worksheet для копіювання
    ws = wb.copy_worksheet(source_ws)
    ws.title = "Видатки (формули)"
    
    # Змінюємо формат відображення чисел для всіх числових комірок
    for row in range(4, 13):  # Рядки з даними включаючи мінімум, максимум, середнє
        for col in range(2, 14):  # Стовпці з числами
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                # Змінюємо формат на відображення з двома знаками після коми
                if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith('=')):
                    cell.number_format = '0.00'
    
    # Встановлюємо режим відображення формул
    # Примітка: коли showFormulas = True, Excel показує формули як текст замість значень
    # Це стандартна поведінка Excel. Для перегляду значень можна вимкнути режим (Ctrl+`)
    # Але згідно з завданням, потрібно встановити режим відображення формул
    # Тому залишаємо його увімкненим - користувач може вимкнути його в Excel для перегляду значень
    ws.sheet_view.showFormulas = True

def create_monthly_expenses_sheet(wb):
    ws = wb.create_sheet("Видатки на місяць")
    
    # Заголовок
    ws['A1'] = "ВИДАТКИ НА МІСЯЦЬ (ЖОВТЕНЬ)"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Заголовки таблиці
    headers = ['Дата', 'Продукти', 'Транспорт', 'Комунальні', 'Розваги', 'Сума за день']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        # Без кольорів - базовий вигляд
        # cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Генеруємо робочі дні жовтня 2024
    start_date = datetime(2024, 10, 1)
    row = 4
    import random
    random.seed(42)
    
    current_date = start_date
    while current_date.month == 10:
        # Пропускаємо вихідні (субота=5, неділя=6)
        if current_date.weekday() < 5:  # Понеділок-П'ятниця
            ws.cell(row=row, column=1, value=current_date.strftime("%d.%m.%Y"))
            ws.cell(row=row, column=1).number_format = 'DD.MM.YYYY'
            
            # Генеруємо витрати
            for col in range(2, 6):
                amount = random.randint(0, 500)
                ws.cell(row=row, column=col, value=amount)
                ws.cell(row=row, column=col).number_format = '#,##0'
            
            # Сума за день
            sum_formula = f"=SUM(B{row}:E{row})"
            ws.cell(row=row, column=6, value=sum_formula)
            ws.cell(row=row, column=6).number_format = '#,##0'
            
            # Додаємо межі
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            row += 1
        
        current_date += timedelta(days=1)
    
    # Налаштування ширини стовпців
    ws.column_dimensions['A'].width = 15
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Графік динаміки витрат
    chart = LineChart()
    chart.title = "Динаміка витрат протягом місяця"
    chart.style = 13
    chart.y_axis.title = 'Сума (грн)'
    chart.x_axis.title = 'Дата'
    
    last_row = row - 1
    data = Reference(ws, min_col=6, min_row=3, max_col=6, max_row=last_row)
    cats = Reference(ws, min_col=1, min_row=4, max_col=1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws.add_chart(chart, "A30")

def create_price_sheet(wb):
    ws = wb.create_sheet("Прайс")
    
    # Заголовок
    ws['A1'] = "ПРАЙС-ЛИСТ ФІРМИ «Compservise»"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Курс долара
    ws['B3'] = "Курс долара:"
    ws['C3'] = 37.5
    ws['C3'].number_format = '#,##0.00'
    
    # Заголовки таблиці
    headers = ['Програмне забезпечення', 'Вартість (USD)', 'Вартість (UAH)']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        # Без кольорів - базовий вигляд
        # cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Програмне забезпечення
    software = [
        ('Windows 11 Pro', 199),
        ('Microsoft Office 365', 99),
        ('Adobe Photoshop', 299),
        ('Visual Studio Code', 0),
        ('AutoCAD', 1299),
        ('Norton Antivirus', 49),
        ('WinRAR', 29),
        ('Steam', 0)
    ]
    
    row = 6
    for name, price_usd in software:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=price_usd)
        ws.cell(row=row, column=2).number_format = '#,##0'
        
        # Формула з абсолютною адресацією курсу
        formula = f"=B{row}*$C$3"
        ws.cell(row=row, column=3, value=formula)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        
        # Додаємо межі
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row += 1
    
    # Налаштування ширини стовпців
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    
    # Діаграма вартості у доларах
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Вартість програмного забезпечення (USD)"
    chart.y_axis.title = 'Вартість (USD)'
    chart.x_axis.title = 'Програмне забезпечення'
    
    data = Reference(ws, min_col=2, min_row=5, max_col=2, max_row=row-1)
    cats = Reference(ws, min_col=1, min_row=6, max_col=1, max_row=row-1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws.add_chart(chart, "E5")

def create_function_sheet(wb):
    ws = wb.create_sheet("Функція")
    
    # Заголовок
    ws['A1'] = "ТАБУЛЯЦІЯ ФУНКЦІЇ y = x² + sin(x)"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:C1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Параметри
    ws['A3'] = "Проміжок:"
    ws['B3'] = "[-1; 1]"
    ws['A4'] = "Крок:"
    ws['B4'] = "0.1"
    
    # Заголовки таблиці
    headers = ['x', 'y = x² + sin(x)']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        # Без кольорів - базовий вигляд
        # cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Табулюємо функцію
    x_start = -1.0
    x_end = 1.0
    step = 0.1
    
    row = 7
    x = x_start
    while x <= x_end + 0.0001:  # Додаємо невелику похибку для float
        ws.cell(row=row, column=1, value=x)
        ws.cell(row=row, column=1).number_format = '0.0'
        
        # Формула y = x² + sin(x)
        formula = f"=A{row}^2+SIN(A{row})"
        ws.cell(row=row, column=2, value=formula)
        ws.cell(row=row, column=2).number_format = '0.0000'
        
        # Додаємо межі
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        x += step
        row += 1
    
    # Налаштування ширини стовпців
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    
    # Графік функції
    chart = LineChart()
    chart.title = "Графік функції y = x² + sin(x)"
    chart.style = 13
    chart.y_axis.title = 'y'
    chart.x_axis.title = 'x'
    
    last_row = row - 1
    data = Reference(ws, min_col=2, min_row=6, max_col=2, max_row=last_row)
    cats = Reference(ws, min_col=1, min_row=7, max_col=1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.smooth = True
    
    ws.add_chart(chart, "D6")

if __name__ == "__main__":
    create_spreadsheet()

