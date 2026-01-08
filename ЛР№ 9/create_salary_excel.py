"""
Лабораторна робота: Електронні таблиці - Зарплата та статистика ІТ
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

def create_salary_excel():
    wb = openpyxl.Workbook()
    
    # Стилі
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    currency_format = '#,##0.00 ₴'
    
    # ============ ЛИСТ 1: Зарплата ============
    ws1 = wb.active
    ws1.title = "Зарплата"
    ws1.sheet_properties.tabColor = "FF6600"  # Оранжевий колір ярлика
    
    # Заголовок таблиці
    ws1.merge_cells('A1:F1')
    ws1['A1'] = "Зарплата працівників в ІТ"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A1'].alignment = center_align
    
    # Заголовки стовпців
    headers = ['№ п/п', 'Ідентифікаційний код', 'Прізвище та ім\'я', 
               'Нарахована зарплата (грн)', 'Податок із зарплати (грн)', 'До видачі (грн)']
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    # Дані працівників (12 осіб)
    employees = [
        (1, '1234567890', 'Коваленко Олександр', 7500),
        (2, '2345678901', 'Шевченко Марія', 12000),
        (3, '3456789012', 'Бондаренко Іван', 25000),
        (4, '4567890123', 'Мельник Анна', 8000),
        (5, '5678901234', 'Ткаченко Петро', 35000),
        (6, '6789012345', 'Кравченко Олена', 15000),
        (7, '7890123456', 'Поліщук Андрій', 18500),
        (8, '8901234567', 'Савченко Юлія', 6500),
        (9, '9012345678', 'Лисенко Максим', 42000),
        (10, '0123456789', 'Гончаренко Тетяна', 9500),
        (11, '1122334455', 'Романенко Дмитро', 28000),
        (12, '2233445566', 'Кузьменко Вікторія', 11000),
    ]
    
    for row_idx, (num, code, name, salary) in enumerate(employees, 3):
        ws1.cell(row=row_idx, column=1, value=num).border = border
        ws1.cell(row=row_idx, column=2, value=code).border = border
        ws1.cell(row=row_idx, column=3, value=name).border = border
        
        salary_cell = ws1.cell(row=row_idx, column=4, value=salary)
        salary_cell.border = border
        salary_cell.number_format = currency_format
        
        # Формула для податку: IF(D<8000, D*10%, IF(D>20000, D*20%, D*15%))
        tax_formula = f'=IF(D{row_idx}<=8000,D{row_idx}*0.1,IF(D{row_idx}>20000,D{row_idx}*0.2,D{row_idx}*0.15))'
        tax_cell = ws1.cell(row=row_idx, column=5, value=tax_formula)
        tax_cell.border = border
        tax_cell.number_format = currency_format
        
        # Формула для "До видачі": Зарплата - Податок
        net_formula = f'=D{row_idx}-E{row_idx}'
        net_cell = ws1.cell(row=row_idx, column=6, value=net_formula)
        net_cell.border = border
        net_cell.number_format = currency_format
    
    # Підсумковий рядок
    total_row = len(employees) + 3
    ws1.cell(row=total_row, column=3, value="РАЗОМ:").font = Font(bold=True)
    ws1.cell(row=total_row, column=4, value=f'=SUM(D3:D{total_row-1})').number_format = currency_format
    ws1.cell(row=total_row, column=5, value=f'=SUM(E3:E{total_row-1})').number_format = currency_format
    ws1.cell(row=total_row, column=6, value=f'=SUM(F3:F{total_row-1})').number_format = currency_format
    
    # Ширина стовпців
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 25
    ws1.column_dimensions['D'].width = 25
    ws1.column_dimensions['E'].width = 25
    ws1.column_dimensions['F'].width = 20
    
    # ============ ЛИСТ 2: Формули ============
    ws2 = wb.create_sheet("Формули")
    ws2.sheet_properties.tabColor = "00B050"  # Зелений
    
    # Копіюємо дані з листа 1
    ws2.merge_cells('A1:F1')
    ws2['A1'] = "Зарплата працівників в ІТ (РЕЖИМ ФОРМУЛ)"
    ws2['A1'].font = Font(bold=True, size=14)
    
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=2, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
    
    for row_idx, (num, code, name, salary) in enumerate(employees, 3):
        ws2.cell(row=row_idx, column=1, value=num)
        ws2.cell(row=row_idx, column=2, value=code)
        ws2.cell(row=row_idx, column=3, value=name)
        ws2.cell(row=row_idx, column=4, value=salary)
        # Показуємо формули як текст
        ws2.cell(row=row_idx, column=5, value=f'=IF(D{row_idx}<=8000,D{row_idx}*0.1,IF(D{row_idx}>20000,D{row_idx}*0.2,D{row_idx}*0.15))')
        ws2.cell(row=row_idx, column=6, value=f'=D{row_idx}-E{row_idx}')
    
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 65
    ws2.column_dimensions['F'].width = 20
    
    # ============ ЛИСТ 3: Статистика ІТ ============
    ws3 = wb.create_sheet("Статистика ІТ")
    ws3.sheet_properties.tabColor = "7030A0"  # Фіолетовий
    
    ws3.merge_cells('A1:E1')
    ws3['A1'] = "Статистика ІТ-галузі в Україні"
    ws3['A1'].font = Font(bold=True, size=14)
    ws3['A1'].alignment = center_align
    
    stat_headers = ['№ п/п', 'Область', 'Кількість ФОП-чоловіки (%)', 
                    'Кількість ФОП-жінки (%)', 'Сума сплачених податків ФОП-ІТ (грн)']
    
    for col, header in enumerate(stat_headers, 1):
        cell = ws3.cell(row=2, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Дані статистики по областях
    stats_data = [
        (1, 'Київська', 45, 55, 1250000),
        (2, 'Львівська', 38, 62, 890000),
        (3, 'Харківська', 42, 58, 780000),
        (4, 'Одеська', 35, 65, 620000),
        (5, 'Дніпропетровська', 40, 60, 950000),
        (6, 'Запорізька', 28, 72, 380000),
        (7, 'Вінницька', 22, 78, 290000),
        (8, 'Полтавська', 31, 69, 420000),
        (9, 'Чернігівська', 25, 75, 310000),
        (10, 'Черкаська', 33, 67, 450000),
        (11, 'Чернівецька', 21, 79, 280000),
        (12, 'Волинська', 19, 81, 220000),
        (13, 'Рівненська', 24, 76, 340000),
        (14, 'Тернопільська', 27, 73, 360000),
        (15, 'Івано-Франківська', 29, 71, 410000),
        (16, 'Закарпатська', 18, 82, 190000),
        (17, 'Хмельницька', 26, 74, 350000),
        (18, 'Житомирська', 23, 77, 320000),
        (19, 'Сумська', 30, 70, 390000),
        (20, 'Миколаївська', 32, 68, 430000),
    ]
    
    for row_idx, (num, region, male_pct, female_pct, taxes) in enumerate(stats_data, 3):
        ws3.cell(row=row_idx, column=1, value=num).border = border
        ws3.cell(row=row_idx, column=2, value=region).border = border
        ws3.cell(row=row_idx, column=3, value=male_pct).border = border
        ws3.cell(row=row_idx, column=4, value=female_pct).border = border
        tax_cell = ws3.cell(row=row_idx, column=5, value=taxes)
        tax_cell.border = border
        tax_cell.number_format = '#,##0 ₴'
    
    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 22
    ws3.column_dimensions['C'].width = 25
    ws3.column_dimensions['D'].width = 22
    ws3.column_dimensions['E'].width = 32
    ws3.row_dimensions[2].height = 40
    
    # ============ ЛИСТ 4: Автофільтр ============
    ws4 = wb.create_sheet("Автофільтр")
    ws4.sheet_properties.tabColor = "FFC000"  # Жовтий
    
    ws4.merge_cells('A1:E1')
    ws4['A1'] = "Статистика ІТ-галузі в Україні (Автофільтр)"
    ws4['A1'].font = Font(bold=True, size=14)
    ws4['A1'].alignment = center_align
    
    for col, header in enumerate(stat_headers, 1):
        cell = ws4.cell(row=2, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row_idx, (num, region, male_pct, female_pct, taxes) in enumerate(stats_data, 3):
        ws4.cell(row=row_idx, column=1, value=num).border = border
        ws4.cell(row=row_idx, column=2, value=region).border = border
        ws4.cell(row=row_idx, column=3, value=male_pct).border = border
        ws4.cell(row=row_idx, column=4, value=female_pct).border = border
        tax_cell = ws4.cell(row=row_idx, column=5, value=taxes)
        tax_cell.border = border
        tax_cell.number_format = '#,##0 ₴'
    
    # Додаємо автофільтр
    ws4.auto_filter.ref = f"A2:E{len(stats_data)+2}"
    
    ws4.column_dimensions['A'].width = 8
    ws4.column_dimensions['B'].width = 22
    ws4.column_dimensions['C'].width = 25
    ws4.column_dimensions['D'].width = 22
    ws4.column_dimensions['E'].width = 32
    ws4.row_dimensions[2].height = 40
    
    # Додаємо інструкції для фільтрів
    ws4['G2'] = "ІНСТРУКЦІЇ ДЛЯ ФІЛЬТРАЦІЇ:"
    ws4['G2'].font = Font(bold=True, color="FF0000")
    ws4['G3'] = "11. Фільтр по 'Область' → Текстові фільтри → Починається з 'Ч'"
    ws4['G4'] = "12. Фільтр по 'Сума податків' → Числові фільтри → Більше ніж 500000"
    ws4['G5'] = "13. Фільтр по 'ФОП-чоловіки' → Числові фільтри → Між 20 і 40"
    ws4['G6'] = "14. Дані → Очистити фільтр (показати всі)"
    
    # ============ ЛИСТ 5: Розширений фільтр ============
    ws5 = wb.create_sheet("Розширений фільтр")
    ws5.sheet_properties.tabColor = "00B0F0"  # Блакитний
    
    ws5.merge_cells('A1:F1')
    ws5['A1'] = "Діяльність працівників фірми ІТ-Market@"
    ws5['A1'].font = Font(bold=True, size=14)
    ws5['A1'].alignment = center_align
    
    market_headers = ['№ п/п', 'Прізвище', 'Ім\'я', 'Посада', 'Відділ', 'Внесок ($)']
    
    for col, header in enumerate(market_headers, 1):
        cell = ws5.cell(row=2, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    # Дані працівників фірми
    market_employees = [
        (1, 'Василенко', 'Олександр', 'Розробник', 'Backend', 750),
        (2, 'Бондаренко', 'Марія', 'Тестувальник', 'QA', 520),
        (3, 'Петренко', 'Іван', 'Менеджер', 'Продажі', 1200),
        (4, 'Василенко', 'Анна', 'Дизайнер', 'UX/UI', 680),
        (5, 'Коваленко', 'Олександр', 'DevOps', 'Інфраструктура', 890),
        (6, 'Бабенко', 'Петро', 'Розробник', 'Frontend', 820),
        (7, 'Приходько', 'Олена', 'Аналітик', 'Аналітика', 650),
        (8, 'Василенко', 'Максим', 'Розробник', 'Mobile', 780),
        (9, 'Павленко', 'Юлія', 'HR', 'Персонал', 450),
        (10, 'Бойко', 'Олександр', 'Архітектор', 'Backend', 1100),
        (11, 'Поліщук', 'Тетяна', 'Тестувальник', 'QA', 580),
        (12, 'Барановський', 'Дмитро', 'Розробник', 'Backend', 720),
        (13, 'Кузьменко', 'Олександр', 'Менеджер', 'Проєкти', 950),
        (14, 'Притула', 'Вікторія', 'Дизайнер', 'UX/UI', 610),
        (15, 'Бондар', 'Сергій', 'Розробник', 'Frontend', 830),
    ]
    
    for row_idx, (num, surname, name, position, dept, contribution) in enumerate(market_employees, 3):
        ws5.cell(row=row_idx, column=1, value=num).border = border
        ws5.cell(row=row_idx, column=2, value=surname).border = border
        ws5.cell(row=row_idx, column=3, value=name).border = border
        ws5.cell(row=row_idx, column=4, value=position).border = border
        ws5.cell(row=row_idx, column=5, value=dept).border = border
        contr_cell = ws5.cell(row=row_idx, column=6, value=contribution)
        contr_cell.border = border
        contr_cell.number_format = '$#,##0'
    
    ws5.column_dimensions['A'].width = 8
    ws5.column_dimensions['B'].width = 18
    ws5.column_dimensions['C'].width = 15
    ws5.column_dimensions['D'].width = 15
    ws5.column_dimensions['E'].width = 18
    ws5.column_dimensions['F'].width = 12
    
    # Область критеріїв для розширеного фільтра
    ws5['H2'] = "КРИТЕРІЇ ДЛЯ РОЗШИРЕНОГО ФІЛЬТРА:"
    ws5['H2'].font = Font(bold=True, color="FF0000")
    
    ws5['H4'] = "Критерій 16: Ім'я"
    ws5['I4'] = "Олександр"
    ws5['H4'].font = Font(bold=True)
    ws5['H5'] = "(найпоширеніше ім'я - 5 осіб)"
    
    ws5['H7'] = "Критерій 17: Прізвище"
    ws5['I7'] = "Василенко"
    ws5['H7'].font = Font(bold=True)
    
    ws5['H9'] = "Критерій 18: Прізвище"
    ws5['I9'] = "Б*"
    ws5['H10'] = "Прізвище"
    ws5['I10'] = "П*"
    ws5['H9'].font = Font(bold=True)
    ws5['H11'] = "(починаються на Б або П)"
    
    ws5['H13'] = "Критерій 19: Внесок ($)"
    ws5['I13'] = ">=500"
    ws5['H14'] = "Внесок ($)"
    ws5['I14'] = "<=900"
    ws5['H13'].font = Font(bold=True)
    ws5['H15'] = "(від 500$ до 900$)"
    
    # Зберігаємо файл
    filename = "/Users/maksymyanisiv/test for cursor/Yanisiv_LR.xlsx"
    wb.save(filename)
    print(f"✅ Файл створено: {filename}")
    print("\n📋 Структура файлу:")
    print("   1. Лист 'Зарплата' - таблиця з формулами податків (оранжевий ярлик)")
    print("   2. Лист 'Формули' - копія з відображенням формул (зелений ярлик)")
    print("   3. Лист 'Статистика ІТ' - дані по областях (фіолетовий ярлик)")
    print("   4. Лист 'Автофільтр' - з інструкціями для фільтрації (жовтий ярлик)")
    print("   5. Лист 'Розширений фільтр' - ІТ-Market@ з критеріями (блакитний ярлик)")
    print("\n📌 Формула податку:")
    print("   • Зарплата ≤ 8000 грн → 10%")
    print("   • Зарплата > 20000 грн → 20%")
    print("   • Інші випадки → 15%")

if __name__ == "__main__":
    create_salary_excel()

