import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Середній бал студентів"

students = [
    ("Гордюк М.В.", "Інформатика"),
    ("Гриценко А.П.", "Економіка"),
    ("Губська Т.А.", "Комп'ютерні науки"),
    ("Дзюба Д.Г.", "Економіка"),
    ("Євдокименко Ю.Ф.", "Економіка"),
    ("Кукса Ю.О.", "Фінанси"),
    ("Мороз О.М.", "Комп'ютерні науки"),
    ("Патлах О.О.", "Комп'ютерні науки"),
    ("Петренко Л.Г.", "Інформатика"),
    ("Роботько О.М.", "Інформатика"),
    ("Плацинда М.М.", "Інформатика"),
]

headers = [
    "№",
    "Прізвище",
    "Спеціальність",
    "Предмет 1",
    "Предмет 2",
    "Предмет 3",
    "Предмет 4",
    "Предмет 5",
    "Середній бал"
]

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for row, (name, specialty) in enumerate(students, 2):
    ws.cell(row=row, column=1, value=row-1).alignment = center_align
    ws.cell(row=row, column=1).border = thin_border
    
    ws.cell(row=row, column=2, value=name).alignment = left_align
    ws.cell(row=row, column=2).border = thin_border
    
    ws.cell(row=row, column=3, value=specialty).alignment = left_align
    ws.cell(row=row, column=3).border = thin_border
    
    for col in range(4, 9):
        grade = round(random.uniform(0, 100), 2)
        cell = ws.cell(row=row, column=col, value=grade)
        cell.number_format = '0.00'
        cell.alignment = center_align
        cell.border = thin_border
    
    avg_cell = ws.cell(row=row, column=9)
    avg_cell.value = f"=AVERAGE(D{row}:H{row})"
    avg_cell.number_format = '0.00'
    avg_cell.alignment = center_align
    avg_cell.border = thin_border

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 20
for col in range(4, 9):
    ws.column_dimensions[get_column_letter(col)].width = 12
ws.column_dimensions['I'].width = 14

ws.row_dimensions[1].height = 30

output_file = "/Users/maksymyanisiv/test for cursor/Студенти_середній_бал.xlsx"
wb.save(output_file)
print(f"Файл створено: {output_file}")
